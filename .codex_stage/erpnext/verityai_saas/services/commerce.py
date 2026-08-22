from html import escape
from io import BytesIO
import hashlib
import hmac
import time
from urllib.parse import urlencode

import frappe
from frappe.utils import add_to_date, flt, get_datetime, get_url, getdate, now_datetime, nowdate, validate_email_address
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CUSTOMER_FIELDS = ["name", "customer_name", "customer_type", "email", "phone", "tax_id", "address", "city", "country", "notes", "status", "external_system", "external_id", "last_synced_on", "creation", "modified"]
PRODUCT_FIELDS = ["name", "item_code", "item_name", "description", "item_group", "stock_uom", "is_stock_item", "standard_rate", "currency", "active", "external_system", "external_id", "last_synced_on", "creation", "modified"]
PRODUCT_IMPORT_HEADERS = [
	"Item Code", "Item Name", "Description", "Category", "Unit of Measure",
	"Selling Price", "Currency", "Item Type", "Active",
]
PRICE_FIELDS = ["name", "product", "price_list", "currency", "rate", "valid_from", "valid_upto", "active", "creation", "modified"]
QUOTE_FIELDS = ["name", "customer", "customer_name", "customer_email", "transaction_date", "valid_till", "price_list", "currency", "status", "subtotal", "discount_amount", "tax_rate", "tax_amount", "total", "external_system", "external_id", "sync_status", "last_synced_on", "creation", "modified"]
QUOTE_TRANSITIONS = {
	"Draft": {"Pending Approval", "Cancelled"},
	"Pending Approval": {"Approved", "Rejected", "Cancelled"},
	"Approved": {"Sent", "Cancelled"},
	"Sent": {"Accepted", "Rejected", "Expired", "Cancelled"},
}
OPPORTUNITY_TRANSITIONS = {
	"New": {"Qualified", "Lost"},
	"Qualified": {"Proposal", "Lost"},
	"Proposal": {"Negotiation", "Won", "Lost"},
	"Negotiation": {"Won", "Lost"},
}


def _text(value, label, required=False, maximum=140):
	value = str(value or "").strip()
	if required and not value:
		frappe.throw(f"{label} is required.", frappe.ValidationError)
	if len(value) > maximum:
		frappe.throw(f"{label} must not exceed {maximum} characters.", frappe.ValidationError)
	return value


def _limit(value, maximum=200):
	return min(max(int(value or 50), 1), maximum)


def _scoped_name(doctype, workspace, name, label):
	if not name or not frappe.db.exists(doctype, {"name": name, "workspace": workspace}):
		frappe.throw(f"{label} was not found in this workspace.", frappe.DoesNotExistError)
	return name


def _workspace_currency(workspace):
	return frappe.db.get_value("VerityAI Workspace", workspace, "currency") or "USD"


def list_customers(workspace, search=None, status=None, limit=100):
	filters = {"workspace": workspace}
	if status:
		filters["status"] = status
	search = _text(search, "Search", maximum=140)
	or_filters = None
	if search:
		pattern = f"%{search}%"
		or_filters = {"customer_name": ["like", pattern], "email": ["like", pattern], "phone": ["like", pattern]}
	return frappe.get_all("VerityAI Customer", filters=filters, or_filters=or_filters, fields=CUSTOMER_FIELDS, order_by="customer_name asc", limit_page_length=_limit(limit))


def save_customer(workspace, values, customer=None):
	values = values or {}
	name = _text(values.get("customer_name"), "Customer name", required=True)
	email = _text(values.get("email"), "Email", maximum=254)
	if email:
		validate_email_address(email, throw=True)
	duplicate_filters = {"workspace": workspace, "customer_name": name}
	if customer:
		duplicate_filters["name"] = ["!=", customer]
	if frappe.db.exists("VerityAI Customer", duplicate_filters):
		frappe.throw("A customer with this name already exists in the workspace.", frappe.DuplicateEntryError)
	if customer:
		doc = frappe.get_doc("VerityAI Customer", _scoped_name("VerityAI Customer", workspace, customer, "Customer"))
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Customer", "workspace": workspace})
	doc.update({
		"customer_name": name,
		"customer_type": values.get("customer_type") if values.get("customer_type") in {"Company", "Individual"} else "Company",
		"email": email or None,
		"phone": _text(values.get("phone"), "Phone", maximum=40) or None,
		"tax_id": _text(values.get("tax_id"), "Tax ID", maximum=80) or None,
		"address": _text(values.get("address"), "Address", maximum=1000) or None,
		"city": _text(values.get("city"), "City", maximum=140) or None,
		"country": values.get("country") or None,
		"notes": _text(values.get("notes"), "Notes", maximum=2000) or None,
		"status": values.get("status") if values.get("status") in {"Active", "Disabled"} else "Active",
	})
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return frappe.db.get_value("VerityAI Customer", doc.name, CUSTOMER_FIELDS, as_dict=True)


def delete_customer(workspace, customer):
	name = _scoped_name("VerityAI Customer", workspace, customer, "Customer")
	if frappe.db.exists("VerityAI Quotation", {"workspace": workspace, "customer": name}):
		frappe.throw("Customers referenced by quotations cannot be deleted. Disable the customer instead.", frappe.ValidationError)
	frappe.delete_doc("VerityAI Customer", name, ignore_permissions=True)
	return {"deleted": name}


def list_products(workspace, search=None, active=None, limit=100):
	filters = {"workspace": workspace}
	if active not in (None, ""):
		filters["active"] = int(bool(int(active)))
	search = _text(search, "Search", maximum=140)
	or_filters = None
	if search:
		pattern = f"%{search}%"
		or_filters = {"item_code": ["like", pattern], "item_name": ["like", pattern], "description": ["like", pattern]}
	return frappe.get_all("VerityAI Product", filters=filters, or_filters=or_filters, fields=PRODUCT_FIELDS, order_by="item_name asc", limit_page_length=_limit(limit))


def save_product(workspace, values, product=None):
	values = values or {}
	item_code = _text(values.get("item_code"), "Item code", required=True).upper()
	item_name = _text(values.get("item_name"), "Item name", required=True)
	duplicate_filters = {"workspace": workspace, "item_code": item_code}
	if product:
		duplicate_filters["name"] = ["!=", product]
	if frappe.db.exists("VerityAI Product", duplicate_filters):
		frappe.throw("This item code already exists in the workspace.", frappe.DuplicateEntryError)
	if product:
		doc = frappe.get_doc("VerityAI Product", _scoped_name("VerityAI Product", workspace, product, "Product"))
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Product", "workspace": workspace})
	rate = flt(values.get("standard_rate"), 2)
	if rate < 0:
		frappe.throw("Standard rate cannot be negative.", frappe.ValidationError)
	doc.update({
		"item_code": item_code,
		"item_name": item_name,
		"description": _text(values.get("description"), "Description", maximum=10000) or None,
		"item_group": _text(values.get("item_group"), "Item group", maximum=140) or "Services",
		"stock_uom": _text(values.get("stock_uom"), "Unit of measure", maximum=40) or "Unit",
		"is_stock_item": int(bool(int(values.get("is_stock_item") or 0))),
		"standard_rate": rate,
		"currency": values.get("currency") or _workspace_currency(workspace),
		"active": int(bool(int(values.get("active", 1)))),
	})
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return frappe.db.get_value("VerityAI Product", doc.name, PRODUCT_FIELDS, as_dict=True)


def product_import_template():
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "Products"
	sheet.append(PRODUCT_IMPORT_HEADERS)
	sheet.append(["CONSULT-01", "Business consultation", "One hour consultation", "Services", "Hour", 50, "USD", "Service", "Yes"])
	header_fill = PatternFill("solid", fgColor="1F64C8")
	for cell in sheet[1]:
		cell.fill = header_fill
		cell.font = Font(color="FFFFFF", bold=True)
		cell.alignment = Alignment(vertical="center")
	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = f"A1:I2"
	widths = [18, 28, 42, 20, 18, 16, 12, 14, 12]
	for index, width in enumerate(widths, 1):
		sheet.column_dimensions[chr(64 + index)].width = width
	instructions = workbook.create_sheet("Instructions")
	instructions.append(["Product import"])
	instructions.append(["Complete the Products sheet without changing the column headings."])
	instructions.append(["Item Code and Item Name are required. Item Type accepts Product or Service. Active accepts Yes or No."])
	instructions.append(["Remove the example row before importing your own catalogue."])
	instructions["A1"].font = Font(bold=True, size=14, color="1F64C8")
	instructions.column_dimensions["A"].width = 105
	stream = BytesIO()
	workbook.save(stream)
	return stream.getvalue()


def product_export(workspace):
	workbook = load_workbook(BytesIO(product_import_template()))
	sheet = workbook["Products"]
	sheet.delete_rows(2, sheet.max_row - 1)
	for product in list_products(workspace, limit=1000):
		sheet.append([
			product.item_code,
			product.item_name,
			product.description or "",
			product.item_group or "Services",
			product.stock_uom or "Unit",
			flt(product.standard_rate, 2),
			product.currency or _workspace_currency(workspace),
			"Product" if product.is_stock_item else "Service",
			"Yes" if product.active else "No",
		])
	sheet.auto_filter.ref = f"A1:I{max(sheet.max_row, 1)}"
	stream = BytesIO()
	workbook.save(stream)
	return stream.getvalue()


def import_products(workspace, content, update_existing=False):
	try:
		workbook = load_workbook(BytesIO(content), read_only=True, data_only=False, keep_links=False)
	except Exception:
		frappe.throw("The uploaded file is not a valid Excel workbook.", frappe.ValidationError)
	if "Products" not in workbook.sheetnames:
		frappe.throw("The workbook must contain a Products sheet.", frappe.ValidationError)
	sheet = workbook["Products"]
	if sheet.max_row > 1001:
		frappe.throw("A single workbook can contain at most 1,000 products.", frappe.ValidationError)
	headers = [str(cell.value or "").strip() for cell in sheet[1]]
	if headers != PRODUCT_IMPORT_HEADERS:
		frappe.throw("The Products sheet columns do not match the current template.", frappe.ValidationError)
	rows = []
	seen_codes = set()
	for row_number, cells in enumerate(sheet.iter_rows(min_row=2, max_col=len(PRODUCT_IMPORT_HEADERS)), 2):
		if any(cell.data_type == "f" for cell in cells):
			frappe.throw(f"Row {row_number} contains a formula. Replace formulas with values before importing.", frappe.ValidationError)
		values = [cell.value for cell in cells]
		if not any(value not in (None, "") for value in values):
			continue
		if len(rows) >= 1000:
			frappe.throw("A single workbook can contain at most 1,000 products.", frappe.ValidationError)
		item_code = _text(values[0], "Item code", required=True).upper()
		item_name = _text(values[1], "Item name", required=True)
		if item_code in seen_codes:
			frappe.throw(f"Item code {item_code} appears more than once in the workbook.", frappe.ValidationError)
		seen_codes.add(item_code)
		item_type = str(values[7] or "Service").strip().lower()
		if item_type not in {"product", "service"}:
			frappe.throw(f"Row {row_number}: Item Type must be Product or Service.", frappe.ValidationError)
		active_value = str(values[8] if values[8] is not None else "Yes").strip().lower()
		if active_value not in {"yes", "no", "true", "false", "1", "0"}:
			frappe.throw(f"Row {row_number}: Active must be Yes or No.", frappe.ValidationError)
		rate = flt(values[5], 2)
		if rate < 0:
			frappe.throw(f"Row {row_number}: Selling Price cannot be negative.", frappe.ValidationError)
		rows.append({
			"row_number": row_number,
			"item_code": item_code,
			"values": {
				"item_code": item_code,
				"item_name": item_name,
				"description": values[2],
				"item_group": values[3] or "Services",
				"stock_uom": values[4] or "Unit",
				"standard_rate": rate,
				"currency": str(values[6] or _workspace_currency(workspace)).strip().upper(),
				"is_stock_item": int(item_type == "product"),
				"active": int(active_value in {"yes", "true", "1"}),
			},
		})
	if not rows:
		frappe.throw("The Products sheet does not contain any product rows.", frappe.ValidationError)
	created = updated = skipped = 0
	for row in rows:
		existing = frappe.db.get_value("VerityAI Product", {"workspace": workspace, "item_code": row["item_code"]}, "name")
		if existing and not update_existing:
			skipped += 1
			continue
		try:
			save_product(workspace, row["values"], product=existing)
		except Exception as error:
			frappe.throw(f"Row {row['row_number']}: {error}", frappe.ValidationError)
		if existing:
			updated += 1
		else:
			created += 1
	return {"created": created, "updated": updated, "skipped": skipped, "total": len(rows)}


def delete_product(workspace, product):
	name = _scoped_name("VerityAI Product", workspace, product, "Product")
	if frappe.db.exists("VerityAI Quotation Item", {"product": name}):
		frappe.throw("Products referenced by quotations cannot be deleted. Disable the product instead.", frappe.ValidationError)
	for price in frappe.get_all("VerityAI Product Price", filters={"workspace": workspace, "product": name}, pluck="name"):
		frappe.delete_doc("VerityAI Product Price", price, ignore_permissions=True)
	frappe.delete_doc("VerityAI Product", name, ignore_permissions=True)
	return {"deleted": name}


def list_prices(workspace, product=None, price_list=None, limit=200):
	filters = {"workspace": workspace}
	if product:
		filters["product"] = _scoped_name("VerityAI Product", workspace, product, "Product")
	if price_list:
		filters["price_list"] = _text(price_list, "Price list", maximum=140)
	return frappe.get_all("VerityAI Product Price", filters=filters, fields=PRICE_FIELDS, order_by="price_list asc, valid_from desc, creation desc", limit_page_length=_limit(limit))


def save_price(workspace, values, price=None):
	values = values or {}
	product = _scoped_name("VerityAI Product", workspace, values.get("product"), "Product")
	price_list = _text(values.get("price_list"), "Price list", required=True)
	currency = values.get("currency") or _workspace_currency(workspace)
	rate = flt(values.get("rate"), 2)
	if rate < 0:
		frappe.throw("Rate cannot be negative.", frappe.ValidationError)
	valid_from = getdate(values.get("valid_from")) if values.get("valid_from") else None
	valid_upto = getdate(values.get("valid_upto")) if values.get("valid_upto") else None
	if valid_from and valid_upto and valid_upto < valid_from:
		frappe.throw("Valid until cannot be before valid from.", frappe.ValidationError)
	if price:
		doc = frappe.get_doc("VerityAI Product Price", _scoped_name("VerityAI Product Price", workspace, price, "Price"))
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Product Price", "workspace": workspace})
	doc.update({"product": product, "price_list": price_list, "currency": currency, "rate": rate, "valid_from": valid_from, "valid_upto": valid_upto, "active": int(bool(int(values.get("active", 1))))})
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return frappe.db.get_value("VerityAI Product Price", doc.name, PRICE_FIELDS, as_dict=True)


def delete_price(workspace, price):
	name = _scoped_name("VerityAI Product Price", workspace, price, "Price")
	frappe.delete_doc("VerityAI Product Price", name, ignore_permissions=True)
	return {"deleted": name}


def _product_rate(workspace, product, price_list, currency, on_date):
	product_defaults = frappe.db.get_value("VerityAI Product", product, ["standard_rate", "currency"], as_dict=True)
	if price_list == "Standard Selling":
		if product_defaults.currency != currency:
			frappe.throw(f"No {currency} price is configured for this product.", frappe.ValidationError)
		return flt(product_defaults.standard_rate, 2)
	rows = frappe.get_all("VerityAI Product Price", filters={"workspace": workspace, "product": product, "price_list": price_list, "currency": currency, "active": 1}, fields=["rate", "valid_from", "valid_upto"], order_by="valid_from desc, creation desc", limit_page_length=100)
	for row in rows:
		if row.valid_from and getdate(row.valid_from) > on_date:
			continue
		if row.valid_upto and getdate(row.valid_upto) < on_date:
			continue
		return flt(row.rate, 2)
	if product_defaults.currency != currency:
		frappe.throw(f"No {currency} price is configured for this product.", frappe.ValidationError)
	return flt(product_defaults.standard_rate, 2)


def _quote_items(workspace, items, price_list, currency, transaction_date):
	if not isinstance(items, (list, tuple)) or not items:
		frappe.throw("At least one quotation item is required.", frappe.ValidationError)
	rows = []
	for value in items[:100]:
		product_name = _scoped_name("VerityAI Product", workspace, value.get("product"), "Product")
		product = frappe.db.get_value("VerityAI Product", product_name, ["item_code", "item_name", "description", "stock_uom", "active"], as_dict=True)
		if not product.active:
			frappe.throw(f"Product {product.item_code} is disabled.", frappe.ValidationError)
		qty = flt(value.get("qty"), 4)
		if qty <= 0:
			frappe.throw("Quotation item quantity must be greater than zero.", frappe.ValidationError)
		discount = flt(value.get("discount_percent"), 2)
		if discount < 0 or discount > 100:
			frappe.throw("Item discount must be between 0 and 100 percent.", frappe.ValidationError)
		rate = _product_rate(workspace, product_name, price_list, currency, transaction_date) if value.get("rate") in (None, "") else flt(value.get("rate"), 2)
		if rate < 0:
			frappe.throw("Quotation item rate cannot be negative.", frappe.ValidationError)
		amount = flt(qty * rate * (1 - discount / 100), 2)
		rows.append({"product": product_name, "item_code": product.item_code, "item_name": product.item_name, "description": _text(value.get("description"), "Item description", maximum=1000) or product.description, "qty": qty, "uom": product.stock_uom, "rate": rate, "discount_percent": discount, "amount": amount})
	return rows


def save_quotation(workspace, values, quotation=None):
	values = values or {}
	customer_name = _scoped_name("VerityAI Customer", workspace, values.get("customer"), "Customer")
	customer = frappe.db.get_value("VerityAI Customer", customer_name, ["customer_name", "email", "status"], as_dict=True)
	if customer.status != "Active":
		frappe.throw("The selected customer is disabled.", frappe.ValidationError)
	transaction_date = getdate(values.get("transaction_date") or nowdate())
	valid_till = getdate(values.get("valid_till")) if values.get("valid_till") else None
	if valid_till and valid_till < transaction_date:
		frappe.throw("Valid until cannot be before the quotation date.", frappe.ValidationError)
	price_list = _text(values.get("price_list"), "Price list", maximum=140) or "Standard Selling"
	currency = values.get("currency") or _workspace_currency(workspace)
	items = _quote_items(workspace, values.get("items"), price_list, currency, transaction_date)
	subtotal = flt(sum(row["amount"] for row in items), 2)
	discount_amount = flt(values.get("discount_amount"), 2)
	if discount_amount < 0 or discount_amount > subtotal:
		frappe.throw("Additional discount must be between zero and the subtotal.", frappe.ValidationError)
	tax_rate = flt(values.get("tax_rate"), 2)
	if tax_rate < 0 or tax_rate > 100:
		frappe.throw("Tax rate must be between 0 and 100 percent.", frappe.ValidationError)
	tax_amount = flt((subtotal - discount_amount) * tax_rate / 100, 2)
	total = flt(subtotal - discount_amount + tax_amount, 2)
	if quotation:
		doc = frappe.get_doc("VerityAI Quotation", _scoped_name("VerityAI Quotation", workspace, quotation, "Quotation"))
		if doc.status not in {"Draft", "Pending Approval"}:
			frappe.throw("Only draft or pending quotations can be edited.", frappe.ValidationError)
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Quotation", "workspace": workspace, "status": "Draft"})
	doc.update({"customer": customer_name, "customer_name": customer.customer_name, "customer_email": customer.email, "transaction_date": transaction_date, "valid_till": valid_till, "price_list": price_list, "currency": currency, "subtotal": subtotal, "discount_amount": discount_amount, "tax_rate": tax_rate, "tax_amount": tax_amount, "total": total, "notes": _text(values.get("notes"), "Notes", maximum=10000) or None})
	doc.set("items", items)
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return get_quotation(workspace, doc.name)


def list_quotations(workspace, status=None, customer=None, limit=100):
	filters = {"workspace": workspace}
	if status:
		filters["status"] = status
	if customer:
		filters["customer"] = _scoped_name("VerityAI Customer", workspace, customer, "Customer")
	return frappe.get_all("VerityAI Quotation", filters=filters, fields=QUOTE_FIELDS, order_by="transaction_date desc, creation desc", limit_page_length=_limit(limit))


def get_quotation(workspace, quotation):
	doc = frappe.get_doc("VerityAI Quotation", _scoped_name("VerityAI Quotation", workspace, quotation, "Quotation"))
	data = {field: doc.get(field) for field in QUOTE_FIELDS}
	data["notes"] = doc.notes
	data["items"] = [{key: row.get(key) for key in ("name", "product", "item_code", "item_name", "description", "qty", "uom", "rate", "discount_percent", "amount")} for row in doc.items]
	return frappe._dict(data)


def set_quotation_status(workspace, quotation, status):
	doc = frappe.get_doc("VerityAI Quotation", _scoped_name("VerityAI Quotation", workspace, quotation, "Quotation"))
	status = _text(status, "Status", required=True)
	if status not in QUOTE_TRANSITIONS.get(doc.status, set()):
		frappe.throw(f"Quotation cannot move from {doc.status} to {status}.", frappe.ValidationError)
	doc.status = status
	doc.save(ignore_permissions=True)
	if status == "Approved":
		from verityai_saas.services import erpnext

		erpnext.auto_sync_quotation(workspace, doc.name)
	return get_quotation(workspace, doc.name)


def _quote_signing_secret():
	secret = (getattr(frappe.local, "conf", {}) or {}).get("encryption_key")
	if not secret:
		frappe.throw("Quotation downloads are unavailable until the site encryption key is configured.", frappe.ValidationError)
	return str(secret).encode("utf-8")


def public_quotation_token(workspace, quotation, expires=None):
	expires = int(expires or add_to_date(now_datetime(), days=7).timestamp())
	payload = f"{workspace}:{quotation}:{expires}".encode("utf-8")
	signature = hmac.new(_quote_signing_secret(), payload, hashlib.sha256).hexdigest()
	return f"{expires}.{signature}"


def verify_public_quotation_token(workspace, quotation, token):
	try:
		expires_text, supplied = str(token or "").split(".", 1)
		expires = int(expires_text)
	except (TypeError, ValueError):
		return False
	if expires < int(time.time()):
		return False
	expected = public_quotation_token(workspace, quotation, expires=expires).split(".", 1)[1]
	return hmac.compare_digest(supplied, expected)


def public_quotation_url(workspace, quotation):
	token = public_quotation_token(workspace, quotation)
	query = urlencode({"workspace": workspace, "quotation": quotation, "token": token})
	return f"{get_url().rstrip('/')}/api/method/verityai_saas.api.commerce.download_public_quotation?{query}"


def approve_and_send_quotation(workspace, quotation):
	quote = get_quotation(workspace, quotation)
	if quote.status == "Pending Approval":
		quote = set_quotation_status(workspace, quotation, "Approved")
	elif quote.status not in {"Approved", "Sent"}:
		frappe.throw("Only a pending quotation can be approved.", frappe.ValidationError)
	pdf_url = public_quotation_url(workspace, quotation)
	delivery = "Not sent"
	if quote.customer_email:
		business_name = frappe.db.get_value("VerityAI Workspace", workspace, "business_name") or "VerityAI"
		frappe.sendmail(
			recipients=[quote.customer_email],
			subject=f"Quotation {quote.name} from {business_name}",
			message=(
				f"<p>Hello {escape(quote.customer_name)},</p>"
				f"<p>Your quotation from {escape(business_name)} is ready.</p>"
				f"<p><a href=\"{escape(pdf_url)}\">Download quotation {escape(quote.name)}</a></p>"
				"<p>Please reply to this email if you would like any clarification.</p>"
			),
			reference_doctype="VerityAI Quotation",
			reference_name=quote.name,
		)
		delivery = "Email queued"
		if quote.status == "Approved":
			quote = set_quotation_status(workspace, quotation, "Sent")
	from verityai_saas.services import erpnext

	sync = erpnext.auto_sync_quotation(workspace, quotation)
	return {"quotation": get_quotation(workspace, quotation), "pdf_url": pdf_url, "delivery": delivery, "erpnext_sync": sync}


def _ai_request_product(workspace, requested_label, rate=None):
	label = _text(requested_label, "Requested item", required=True, maximum=240)
	product = frappe.db.get_value(
		"VerityAI Product",
		{"workspace": workspace, "item_code": label.upper(), "active": 1},
		"name",
	)
	if not product:
		product = frappe.db.get_value(
			"VerityAI Product", {"workspace": workspace, "item_name": label, "active": 1}, "name"
		)
	if product:
		return product, None
	placeholder = frappe.db.get_value(
		"VerityAI Product", {"workspace": workspace, "item_code": "AI-CUSTOM-SCOPE"}, "name"
	)
	if not placeholder:
		placeholder = save_product(workspace, {
			"item_code": "AI-CUSTOM-SCOPE", "item_name": "Custom scope",
			"description": "Scope and pricing are confirmed during quotation review.",
			"item_group": "Services", "stock_uom": "Unit", "standard_rate": 0,
			"currency": _workspace_currency(workspace), "active": 1,
		}).name
	return placeholder, label


def _scoped_lead(workspace, lead):
	tenant = frappe.db.get_value("VerityAI Workspace", workspace, "engine_tenant")
	if not lead or not frappe.db.exists("AI Lead", {"name": lead, "tenant": tenant}):
		frappe.throw("Lead was not found in this workspace.", frappe.DoesNotExistError)
	return lead


def _scoped_optional(doctype, workspace, name, label):
	return _scoped_name(doctype, workspace, name, label) if name else None


def _validate_assignee(workspace, user):
	if not user:
		return None
	owner = frappe.db.get_value("VerityAI Workspace", workspace, "owner_user")
	if user != owner and not frappe.db.exists("VerityAI Workspace Member", {"workspace": workspace, "user": user, "status": "Active"}):
		frappe.throw("Assigned user is not an active workspace member.", frappe.ValidationError)
	return user


def convert_lead(workspace, lead, values=None):
	values = values or {}
	lead_name = _scoped_lead(workspace, lead)
	lead_doc = frappe.get_doc("AI Lead", lead_name)
	existing_opportunity = frappe.db.get_value("VerityAI Sales Opportunity", {"workspace": workspace, "lead": lead_name}, ["name", "customer", "stage"], as_dict=True)
	if existing_opportunity:
		return {"lead": lead_name, "customer": existing_opportunity.customer, "opportunity": existing_opportunity.name, "stage": existing_opportunity.stage, "already_converted": True}
	customer = None
	if lead_doc.email:
		customer = frappe.db.get_value("VerityAI Customer", {"workspace": workspace, "email": lead_doc.email}, "name")
	if not customer:
		customer = frappe.db.get_value("VerityAI Customer", {"workspace": workspace, "customer_name": lead_doc.lead_name}, "name")
	if not customer:
		customer = save_customer(workspace, {"customer_name": lead_doc.lead_name, "email": lead_doc.email, "phone": lead_doc.phone, "notes": lead_doc.requirements}).name
	frappe.db.set_value("VerityAI Customer", customer, {"source_lead": lead_name, "converted_on": now_datetime()})
	opportunity = save_opportunity(workspace, {
		"opportunity_name": values.get("opportunity_name") or f"{lead_doc.lead_name} opportunity",
		"lead": lead_name, "customer": customer, "stage": values.get("stage") or "Qualified",
		"amount": values.get("amount"), "currency": values.get("currency"), "probability": values.get("probability", 30),
		"expected_close_date": values.get("expected_close_date"), "source": values.get("source") or lead_doc.source_channel,
		"assigned_to": values.get("assigned_to"), "next_follow_up_on": values.get("next_follow_up_on"), "notes": values.get("notes") or lead_doc.requirements,
	})
	lead_doc.status = "Qualified"
	lead_doc.save(ignore_permissions=True)
	return {"lead": lead_name, "customer": customer, "opportunity": opportunity.name, "stage": opportunity.stage}


def save_opportunity(workspace, values, opportunity=None):
	values = values or {}
	stage = values.get("stage") or "New"
	if stage not in {"New", "Qualified", "Proposal", "Negotiation", "Won", "Lost"}:
		frappe.throw("Unsupported opportunity stage.", frappe.ValidationError)
	lead = _scoped_lead(workspace, values.get("lead")) if values.get("lead") else None
	customer = _scoped_optional("VerityAI Customer", workspace, values.get("customer"), "Customer")
	if not (lead or customer):
		frappe.throw("An opportunity must be linked to a lead or customer.", frappe.ValidationError)
	if opportunity:
		doc = frappe.get_doc("VerityAI Sales Opportunity", _scoped_name("VerityAI Sales Opportunity", workspace, opportunity, "Opportunity"))
		if stage != doc.stage and stage not in OPPORTUNITY_TRANSITIONS.get(doc.stage, set()):
			frappe.throw(f"Opportunity cannot move from {doc.stage} to {stage}.", frappe.ValidationError)
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Sales Opportunity", "workspace": workspace})
	amount = flt(values.get("amount"), 2)
	probability = flt(values.get("probability"), 2)
	if amount < 0 or probability < 0 or probability > 100:
		frappe.throw("Opportunity value must be non-negative and probability must be between 0 and 100.", frappe.ValidationError)
	doc.update({
		"opportunity_name": _text(values.get("opportunity_name"), "Opportunity name", required=True), "lead": lead, "customer": customer,
		"stage": stage, "amount": amount, "currency": values.get("currency") or _workspace_currency(workspace), "probability": probability,
		"expected_close_date": getdate(values.get("expected_close_date")) if values.get("expected_close_date") else None,
		"source": _text(values.get("source"), "Source", maximum=140) or None, "assigned_to": _validate_assignee(workspace, values.get("assigned_to")),
		"next_follow_up_on": get_datetime(values.get("next_follow_up_on")) if values.get("next_follow_up_on") else None,
		"lost_reason": _text(values.get("lost_reason"), "Lost reason", maximum=1000) or None, "notes": _text(values.get("notes"), "Notes", maximum=10000) or None,
	})
	if stage == "Lost" and not doc.lost_reason:
		frappe.throw("Lost reason is required when closing an opportunity as lost.", frappe.ValidationError)
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	if lead and stage in {"Won", "Lost"}:
		frappe.db.set_value("AI Lead", lead, "status", stage)
	if customer and stage == "Won":
		won = frappe.get_all("VerityAI Sales Opportunity", filters={"workspace": workspace, "customer": customer, "stage": "Won"}, pluck="amount")
		frappe.db.set_value("VerityAI Customer", customer, "lifetime_value", flt(sum(flt(value) for value in won), 2))
	return frappe.db.get_value("VerityAI Sales Opportunity", doc.name, ["name", "opportunity_name", "lead", "customer", "stage", "amount", "currency", "probability", "expected_close_date", "source", "assigned_to", "last_contact_on", "next_follow_up_on", "lost_reason", "notes", "creation", "modified"], as_dict=True)


def list_opportunities(workspace, stage=None, assigned_to=None, limit=200):
	filters = {"workspace": workspace}
	if stage:
		filters["stage"] = stage
	if assigned_to:
		filters["assigned_to"] = assigned_to
	rows = frappe.get_all("VerityAI Sales Opportunity", filters=filters, fields=["name", "opportunity_name", "lead", "customer", "stage", "amount", "currency", "probability", "expected_close_date", "assigned_to", "next_follow_up_on", "modified"], order_by="modified desc", limit_page_length=_limit(limit))
	counts = {name: 0 for name in ("New", "Qualified", "Proposal", "Negotiation", "Won", "Lost")}
	values = {name: 0 for name in counts}
	for row in frappe.get_all("VerityAI Sales Opportunity", filters={"workspace": workspace}, fields=["stage", "count(name) as count", "sum(amount) as amount"], group_by="stage"):
		counts[row.stage] = int(row.count or 0)
		values[row.stage] = flt(row.amount, 2)
	tenant = frappe.db.get_value("VerityAI Workspace", workspace, "engine_tenant")
	lead_inbox = frappe.db.count("AI Lead", {"tenant": tenant, "status": "New"}) if tenant else 0
	return {
		"rows": rows, "counts": counts, "values": values, "lead_inbox": lead_inbox,
		"open_value": flt(sum(values[name] for name in ("New", "Qualified", "Proposal", "Negotiation")), 2),
		"won_value": values["Won"],
	}


def set_opportunity_stage(workspace, opportunity, stage, lost_reason=None):
	doc = frappe.get_doc("VerityAI Sales Opportunity", _scoped_name("VerityAI Sales Opportunity", workspace, opportunity, "Opportunity"))
	if stage not in OPPORTUNITY_TRANSITIONS.get(doc.stage, set()):
		frappe.throw(f"Opportunity cannot move from {doc.stage} to {stage}.", frappe.ValidationError)
	if stage == "Lost" and not _text(lost_reason, "Lost reason", maximum=1000):
		frappe.throw("Lost reason is required when closing an opportunity as lost.", frappe.ValidationError)
	doc.stage = stage
	doc.lost_reason = _text(lost_reason, "Lost reason", maximum=1000) or None
	doc.save(ignore_permissions=True)
	if doc.lead and stage in {"Won", "Lost"}:
		frappe.db.set_value("AI Lead", doc.lead, "status", stage)
	if doc.customer and stage == "Won":
		won = frappe.get_all("VerityAI Sales Opportunity", filters={"workspace": workspace, "customer": doc.customer, "stage": "Won"}, pluck="amount")
		frappe.db.set_value("VerityAI Customer", doc.customer, "lifetime_value", flt(sum(flt(value) for value in won), 2))
	return frappe.db.get_value("VerityAI Sales Opportunity", doc.name, ["name", "opportunity_name", "stage", "amount", "currency", "probability", "customer", "lead"], as_dict=True)


def save_appointment(workspace, values, appointment=None):
	values = values or {}
	customer = _scoped_optional("VerityAI Customer", workspace, values.get("customer"), "Customer")
	lead = _scoped_lead(workspace, values.get("lead")) if values.get("lead") else None
	opportunity = _scoped_optional("VerityAI Sales Opportunity", workspace, values.get("opportunity"), "Opportunity")
	if not (customer or lead or opportunity):
		frappe.throw("An appointment must be linked to a customer, lead, or opportunity.", frappe.ValidationError)
	starts_on = get_datetime(values.get("starts_on")) if values.get("starts_on") else None
	ends_on = get_datetime(values.get("ends_on")) if values.get("ends_on") else None
	if not starts_on or (ends_on and ends_on <= starts_on):
		frappe.throw("A valid start time is required and end time must be later.", frappe.ValidationError)
	status = values.get("status") or "Scheduled"
	if status not in {"Scheduled", "Confirmed", "Completed", "Cancelled", "No Show"}:
		frappe.throw("Unsupported appointment status.", frappe.ValidationError)
	if appointment:
		doc = frappe.get_doc("VerityAI Appointment", _scoped_name("VerityAI Appointment", workspace, appointment, "Appointment"))
	else:
		doc = frappe.get_doc({"doctype": "VerityAI Appointment", "workspace": workspace})
	doc.update({"subject": _text(values.get("subject"), "Subject", required=True), "customer": customer, "lead": lead, "opportunity": opportunity, "starts_on": starts_on, "ends_on": ends_on, "timezone": _text(values.get("timezone"), "Timezone", maximum=140) or frappe.db.get_value("VerityAI Workspace", workspace, "timezone"), "mode": values.get("mode") if values.get("mode") in {"Online", "Phone", "Onsite"} else "Online", "location": _text(values.get("location"), "Location", maximum=500) or None, "meeting_url": _text(values.get("meeting_url"), "Meeting URL", maximum=500) or None, "assigned_to": _validate_assignee(workspace, values.get("assigned_to")), "status": status, "notes": _text(values.get("notes"), "Notes", maximum=2000) or None, "outcome": _text(values.get("outcome"), "Outcome", maximum=2000) or None})
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return frappe.db.get_value("VerityAI Appointment", doc.name, ["name", "subject", "customer", "lead", "opportunity", "starts_on", "ends_on", "timezone", "mode", "location", "meeting_url", "assigned_to", "status", "notes", "outcome", "creation", "modified"], as_dict=True)


def list_appointments(workspace, status=None, from_date=None, to_date=None, limit=200):
	filters = {"workspace": workspace}
	if status:
		filters["status"] = status
	if from_date and to_date:
		filters["starts_on"] = ["between", [get_datetime(from_date), get_datetime(to_date)]]
	elif from_date:
		filters["starts_on"] = [">=", get_datetime(from_date)]
	return frappe.get_all("VerityAI Appointment", filters=filters, fields=["name", "subject", "customer", "lead", "opportunity", "starts_on", "ends_on", "timezone", "mode", "location", "meeting_url", "assigned_to", "status", "notes", "outcome", "creation", "modified"], order_by="starts_on asc", limit_page_length=_limit(limit))


def set_appointment_status(workspace, appointment, status, outcome=None):
	doc = frappe.get_doc("VerityAI Appointment", _scoped_name("VerityAI Appointment", workspace, appointment, "Appointment"))
	allowed = {
		"Scheduled": {"Confirmed", "Completed", "Cancelled", "No Show"},
		"Confirmed": {"Completed", "Cancelled", "No Show"},
	}
	if status not in allowed.get(doc.status, set()):
		frappe.throw(f"Appointment cannot move from {doc.status} to {status}.", frappe.ValidationError)
	doc.status = status
	doc.outcome = _text(outcome, "Outcome", maximum=2000) or None
	doc.save(ignore_permissions=True)
	return frappe.db.get_value("VerityAI Appointment", doc.name, ["name", "subject", "status", "outcome", "starts_on"], as_dict=True)


def save_activity(workspace, values):
	values = values or {}
	lead = _scoped_lead(workspace, values.get("lead")) if values.get("lead") else None
	customer = _scoped_optional("VerityAI Customer", workspace, values.get("customer"), "Customer")
	opportunity = _scoped_optional("VerityAI Sales Opportunity", workspace, values.get("opportunity"), "Opportunity")
	appointment = _scoped_optional("VerityAI Appointment", workspace, values.get("appointment"), "Appointment")
	if not (lead or customer or opportunity or appointment):
		frappe.throw("CRM activity must be linked to a lead, customer, opportunity, or appointment.", frappe.ValidationError)
	activity_type = values.get("activity_type") or "Note"
	if activity_type not in {"Call", "Email", "Meeting", "Note", "Follow-up", "Status Change"}:
		frappe.throw("Unsupported activity type.", frappe.ValidationError)
	status = values.get("status") if values.get("status") in {"Open", "Completed", "Cancelled"} else "Open"
	doc = frappe.get_doc({"doctype": "VerityAI CRM Activity", "workspace": workspace, "activity_type": activity_type, "subject": _text(values.get("subject"), "Subject", required=True), "details": _text(values.get("details"), "Details", maximum=5000) or None, "lead": lead, "customer": customer, "opportunity": opportunity, "appointment": appointment, "scheduled_on": get_datetime(values.get("scheduled_on")) if values.get("scheduled_on") else None, "completed_on": now_datetime() if status == "Completed" else None, "assigned_to": _validate_assignee(workspace, values.get("assigned_to")), "status": status}).insert(ignore_permissions=True)
	contact_time = doc.completed_on or now_datetime()
	if customer:
		frappe.db.set_value("VerityAI Customer", customer, "last_contact_on", contact_time)
	if opportunity:
		frappe.db.set_value("VerityAI Sales Opportunity", opportunity, "last_contact_on", contact_time)
	return frappe.db.get_value("VerityAI CRM Activity", doc.name, ["name", "activity_type", "subject", "details", "lead", "customer", "opportunity", "appointment", "scheduled_on", "completed_on", "assigned_to", "status", "creation"], as_dict=True)


def list_activities(workspace, lead=None, customer=None, opportunity=None, status=None, limit=200):
	filters = {"workspace": workspace}
	if lead:
		filters["lead"] = _scoped_lead(workspace, lead)
	if customer:
		filters["customer"] = _scoped_name("VerityAI Customer", workspace, customer, "Customer")
	if opportunity:
		filters["opportunity"] = _scoped_name("VerityAI Sales Opportunity", workspace, opportunity, "Opportunity")
	if status:
		filters["status"] = status
	return frappe.get_all("VerityAI CRM Activity", filters=filters, fields=["name", "activity_type", "subject", "details", "lead", "customer", "opportunity", "appointment", "scheduled_on", "completed_on", "assigned_to", "status", "creation"], order_by="creation desc", limit_page_length=_limit(limit))


def set_activity_status(workspace, activity, status):
	doc = frappe.get_doc("VerityAI CRM Activity", _scoped_name("VerityAI CRM Activity", workspace, activity, "Activity"))
	if doc.status != "Open" or status not in {"Completed", "Cancelled"}:
		frappe.throw(f"Activity cannot move from {doc.status} to {status}.", frappe.ValidationError)
	doc.status = status
	doc.completed_on = now_datetime() if status == "Completed" else None
	doc.save(ignore_permissions=True)
	if status == "Completed":
		if doc.customer:
			frappe.db.set_value("VerityAI Customer", doc.customer, "last_contact_on", doc.completed_on)
		if doc.opportunity:
			frappe.db.set_value("VerityAI Sales Opportunity", doc.opportunity, "last_contact_on", doc.completed_on)
	return frappe.db.get_value("VerityAI CRM Activity", doc.name, ["name", "status", "completed_on"], as_dict=True)


def quotation_html(workspace, quotation):
	doc = get_quotation(workspace, quotation)
	workspace_doc = frappe.get_doc("VerityAI Workspace", workspace)
	currency = escape(str(doc.currency or ""))
	rows = "".join(f"<tr><td>{escape(str(row['item_code']))}</td><td>{escape(str(row['item_name']))}</td><td>{flt(row['qty']):g}</td><td>{currency} {flt(row['rate']):,.2f}</td><td>{currency} {flt(row['amount']):,.2f}</td></tr>" for row in doc["items"])
	return f"""<!doctype html><html><head><meta charset=\"utf-8\"><style>body{{font-family:Arial,sans-serif;color:#172033;margin:40px}}table{{width:100%;border-collapse:collapse}}th,td{{padding:9px;border-bottom:1px solid #ddd;text-align:left}}.total{{font-size:1.2em;font-weight:bold}}</style></head><body><h1>Quotation {escape(doc.name)}</h1><p><strong>{escape(workspace_doc.business_name or workspace_doc.workspace_name)}</strong></p><p>Customer: {escape(doc.customer_name)}<br>Date: {doc.transaction_date}<br>Valid until: {doc.valid_till or 'Not specified'}</p><table><thead><tr><th>Code</th><th>Item</th><th>Qty</th><th>Rate</th><th>Amount</th></tr></thead><tbody>{rows}</tbody></table><p>Subtotal: {currency} {flt(doc.subtotal):,.2f}<br>Discount: {currency} {flt(doc.discount_amount):,.2f}<br>Tax: {currency} {flt(doc.tax_amount):,.2f}</p><p class=\"total\">Total: {currency} {flt(doc.total):,.2f}</p><div>{escape(str(doc.notes or ''))}</div></body></html>"""


def _workspace_for_tenant(tenant_name):
	workspace = frappe.db.get_value("VerityAI Workspace", {"engine_tenant": tenant_name, "status": ["in", ["Trial", "Active"]]}, "name")
	if not workspace:
		return None
	return workspace


def handle_ai_commerce_capabilities(tenant_name):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	return {"handled": True, "native_commerce": True, "catalog": True, "quotations": True, "crm": True}


def handle_ai_catalog_search(tenant_name, query=None, limit=10):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	currency = _workspace_currency(workspace)
	products = list_products(workspace, search=query, active=1, limit=limit)
	rows = []
	for product in products:
		try:
			rate = _product_rate(workspace, product.name, "Standard Selling", currency, getdate(nowdate()))
		except frappe.ValidationError:
			rate = None
		rows.append({
			"item_code": product.item_code,
			"item_name": product.item_name,
			"description": product.description,
			"item_group": product.item_group,
			"uom": product.stock_uom,
			"public_selling_price": rate,
			"currency": currency if rate is not None else None,
		})
	return {"handled": True, "success": True, "products": rows, "count": len(rows)}


def handle_ai_lead_capture(
	tenant_name, lead, name, email=None, phone=None, appointment_requested=None,
	appointment_date=None, appointment_time=None, appointment_mode=None,
	appointment_notes=None, source_channel=None,
):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	_scoped_lead(workspace, lead)
	activity = save_activity(workspace, {
		"activity_type": "Note",
		"subject": "Lead captured by Verity AI",
		"details": f"Lead details captured from {source_channel or 'AI assistant'}.",
		"lead": lead,
		"status": "Completed",
	})
	appointment = None
	if appointment_requested and appointment_date:
		starts_on = get_datetime(f"{appointment_date} {appointment_time or '09:00:00'}")
		appointment = frappe.db.get_value(
			"VerityAI Appointment",
			{"workspace": workspace, "lead": lead, "starts_on": starts_on, "status": ["!=", "Cancelled"]},
			"name",
		)
		if not appointment:
			appointment = save_appointment(workspace, {
				"subject": f"Meeting with {name}",
				"lead": lead,
				"starts_on": starts_on,
				"ends_on": add_to_date(starts_on, minutes=30),
				"mode": appointment_mode or "Online",
				"notes": appointment_notes,
			}).name
	return {"handled": True, "customer": None, "activity": activity.name, "appointment": appointment}


def handle_ai_sales_crm(tenant_name, user, action, reference=None, status=None, values=None, filters=None, limit=50):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	from verityai_saas.services.permissions import require_workspace_permission

	values = values or {}
	filters = filters or {}
	read_actions = {"list_customers", "list_products", "list_quotations", "pipeline_summary", "list_opportunities", "list_appointments", "list_activities"}
	if action in read_actions:
		permission = "view_catalog" if action == "list_products" else "view_quotes" if action == "list_quotations" else "view_customers"
		require_workspace_permission(workspace, permission, user=user)
	else:
		require_workspace_permission(workspace, "manage_customers", user=user)

	if action == "list_customers":
		data = list_customers(workspace, search=filters.get("search"), status=status, limit=limit)
	elif action == "list_products":
		data = list_products(workspace, search=filters.get("search"), active=filters.get("active"), limit=limit)
	elif action == "list_quotations":
		data = list_quotations(workspace, status=status, customer=filters.get("customer"), limit=limit)
	elif action in {"pipeline_summary", "list_opportunities"}:
		data = list_opportunities(workspace, stage=status or filters.get("stage"), assigned_to=filters.get("assigned_to"), limit=limit)
	elif action == "convert_lead":
		data = convert_lead(workspace, reference, values)
	elif action == "set_opportunity_stage":
		data = set_opportunity_stage(workspace, reference, status, lost_reason=values.get("lost_reason"))
	elif action == "list_appointments":
		data = list_appointments(workspace, status=status, from_date=filters.get("from_date"), to_date=filters.get("to_date"), limit=limit)
	elif action == "schedule_appointment":
		data = save_appointment(workspace, values)
	elif action == "set_appointment_status":
		data = set_appointment_status(workspace, reference, status, outcome=values.get("outcome"))
	elif action == "list_activities":
		data = list_activities(workspace, lead=filters.get("lead"), customer=filters.get("customer"), opportunity=filters.get("opportunity"), status=status, limit=limit)
	elif action == "log_activity":
		data = save_activity(workspace, values)
	elif action == "set_activity_status":
		data = set_activity_status(workspace, reference, status)
	else:
		frappe.throw("Unsupported native sales CRM action.", frappe.ValidationError)
	return {"handled": True, "success": True, "action": action, "data": data}


def handle_ai_item_price(tenant_name, item_code):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	item_code = str(item_code or "").strip().upper()
	product = frappe.db.get_value("VerityAI Product", {"workspace": workspace, "item_code": item_code, "active": 1}, "name")
	if not product:
		return {"handled": True, "success": False, "error": f"Product {item_code} was not found in this workspace catalogue."}
	currency = _workspace_currency(workspace)
	rate = _product_rate(workspace, product, "Standard Selling", currency, getdate(nowdate()))
	return {"handled": True, "success": True, "item_code": item_code, "public_selling_price": rate, "currency": currency}


def handle_ai_quotation_request(tenant_name, customer, items, client_whatsapp_number="", client_email=None, notes=None, chat_session=None, source_channel=None):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	filters = {"workspace": workspace, "email": client_email} if client_email else {"workspace": workspace, "customer_name": customer}
	customer_name = frappe.db.get_value("VerityAI Customer", filters, "name")
	if not customer_name:
		customer_name = save_customer(workspace, {"customer_name": customer, "email": client_email, "phone": client_whatsapp_number}).name
	quote_items = []
	for item in items or []:
		requested = str(item.get("item_code") or item.get("item") or item.get("service") or "").strip()
		product, custom_description = _ai_request_product(workspace, requested, rate=item.get("rate"))
		quote_items.append({
			"product": product, "qty": item.get("qty") or 1, "rate": item.get("rate"),
			"discount_percent": item.get("discount_percent") or 0,
			"description": custom_description,
		})
	if not quote_items:
		return {"handled": True, "success": False, "error": "At least one requested product or service is required."}
	quote = save_quotation(workspace, {"customer": customer_name, "items": quote_items, "notes": notes, "price_list": "Standard Selling", "currency": _workspace_currency(workspace)})
	quote = set_quotation_status(workspace, quote.name, "Pending Approval")
	return {"handled": True, "success": True, "quotation_request": quote.name, "quotation": quote.name, "estimated_total": quote.total, "currency": quote.currency, "message": f"Quotation request {quote.name} is pending review."}


def handle_ai_quote_status(tenant_name, quotation_reference=None, customer=None, client_email=None, client_whatsapp_number=None):
	workspace = _workspace_for_tenant(tenant_name)
	if not workspace:
		return None
	if not quotation_reference or not (customer or client_email or client_whatsapp_number):
		return {"handled": True, "success": False, "error": "Please provide the quotation reference and matching customer contact details."}
	if not frappe.db.exists("VerityAI Quotation", {"name": quotation_reference, "workspace": workspace}):
		return {"handled": True, "success": False, "error": "I could not find a matching quotation."}
	quote = get_quotation(workspace, quotation_reference)
	customer_doc = frappe.db.get_value("VerityAI Customer", quote.customer, ["customer_name", "email", "phone"], as_dict=True)
	email_matches = bool(client_email and customer_doc.email and client_email.strip().lower() == customer_doc.email.strip().lower())
	phone_matches = bool(client_whatsapp_number and customer_doc.phone and "".join(filter(str.isdigit, client_whatsapp_number)) == "".join(filter(str.isdigit, customer_doc.phone)))
	name_matches = bool(customer and customer.strip().lower() == customer_doc.customer_name.strip().lower())
	if not (email_matches or phone_matches or name_matches):
		return {"handled": True, "success": False, "error": "The contact details did not match this quotation."}
	pdf_url = public_quotation_url(workspace, quote.name) if quote.status in {"Approved", "Sent", "Accepted"} else None
	return {"handled": True, "success": True, "request": quote.name, "quotation": quote.name, "customer": quote.customer_name, "approval_status": quote.status, "quotation_status": quote.status, "total": quote.total, "currency": quote.currency, "pdf_url": pdf_url}
